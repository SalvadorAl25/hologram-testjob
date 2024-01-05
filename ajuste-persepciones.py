from openpyxl import workbook
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import openpyxl
import pandas as pd
import time

ruta1 = ''
ruta2 = ''
ruta3 = ''

def cargar_archivo1():
    global ruta1 
    ruta1= filedialog.askopenfilename(title="Seleccionar archivo 1", filetypes=(("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")))
    if ruta1:
        dir_path1.config(text=f"Archivo 1 seleccionado: {ruta1}")

def cargar_archivo2():
    global ruta2 
    ruta2 = filedialog.askopenfilename(title="Seleccionar archivo 2", filetypes=(("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")))
    if ruta2:
        dir_path2.config(text=f"Archivo 2 seleccionado: {ruta2}")

def cargar_archivo3():
    global ruta3 
    ruta3 = filedialog.askopenfilename(title="Seleccionar archivo 3", filetypes=(("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")))
    if ruta3:
        dir_path3.config(text=f"Archivo 2 seleccionado: {ruta3}")

def cargar_archivo_excel():
    if ruta1 and ruta2:
        # Cargar el archivo Excel usando pandas
        try:
            t_inicio = time.time()
            # Cargar los archivos de Excel
            libro_ajustes = openpyxl.load_workbook(ruta1)
            libro_original = openpyxl.load_workbook(ruta2)

            hoja_ajuste = libro_ajustes.active
            hoja_original = libro_original.active

            rfc_asign(hoja_ajuste)
            rfc_asign(hoja_original)
           
            transformar(hoja_ajuste,hoja_original)
            codigo_programatico(hoja_original,8)
            print('Guardando Archivo...')
            libro_original.save(ruta2)
            print('Cerrando Archivo...')
            libro_original.close()
            # Mostrar el contenido en la tabla
            mostrar_tabla(hoja_original)
            t_final = time.time()
            tiempo = t_final - t_inicio
            horas = int(tiempo / 3600)
            minutos = int((tiempo % 3600) / 60)
            segundos = int(tiempo % 60)

            mostrar_ventana_emergente(f'Transformación Finalizada, Tiempo de ejecucion {horas}hr:{minutos}min:{segundos}seg')
        except Exception as e:
            # Manejar cualquier error que pueda ocurrir al leer el archivo
            print("Error al cargar el archivo:", e)
    else:
        mostrar_ventana_emergente('Falta alguno de los siguientes archivos: Archivo de ajustes o Archivo Original')

def mostrar_tabla(hoja):
    # Limpiar la tabla si ya existe
    for widget in tabla.winfo_children():
        widget.destroy()

    # Obtener los títulos de las columnas
    columnas = [cell.value for cell in hoja[1]]

    # Crear la tabla usando un widget Treeview de Tkinter
    tabla["columns"] = columnas
    tabla.heading("#0", text="Índice")
    tabla.column("#0", width=50)
    for col in columnas:
        tabla.heading(col, text=col)
        tabla.column(col, width=100)

    # Obtener los datos de las filas y mostrarlos en la tabla
    for idx, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=1):
        tabla.insert("", tk.END, text=str(idx), values=fila)

def mostrar_ventana_emergente(msj):
    ventana_emergente = tk.Toplevel(ventana)
    ventana_emergente.title("Finalizado")
    ventana_emergente.geometry("500x100")
    
    etiqueta = tk.Label(ventana_emergente, text=msj)
    etiqueta.pack(pady=20)
    
    boton_cerrar = tk.Button(ventana_emergente, text="Cerrar ventana", command=ventana_emergente.destroy)
    boton_cerrar.pack(pady=10)

def rfc_asign(hoja):
    rfc_aux = ' '
    
    for fila in hoja.iter_rows(min_row=2):
        if fila[0].value != None:
            rfc_aux = fila[0].value
        else:
            fila[0].value = rfc_aux
         
def buscar_per_rfc(rfc, hoja_ajuste):
    busqueda = []
    for fila in hoja_ajuste.iter_rows(min_row=2):
        if fila[0].value == rfc:
            busqueda.append(fila)
    return busqueda

def transformar(hoja_ajuste,hoja_original):
    rfc_aux = ' '
    busqueda = []
    ban = False
    ind = 0
    tot = hoja_original.max_row - 1
    # Crear un nuevo archivo de Excel
    # este libro es auxiliar, aqui se iran concatenando los registros, para despues sobre escribir la hora destino
    libro_aux = openpyxl.Workbook()
    hoja_aux = libro_aux.active

    print('Transformando...')
    for fila in hoja_original.iter_rows(min_row=2):
        ind = ind + 1
        print(f'Registro: {ind} de {tot}', end='\r')
        if fila[0].value != rfc_aux:
            rfc_aux = fila[0].value
            ban = False
            busqueda = buscar_per_rfc(rfc_aux,hoja_ajuste)
            if len(busqueda) != 0:
                for b in busqueda:
                    hoja_aux.append([i.value for i in b])
                    ban = True
            else:
                hoja_aux.append([j.value for j in fila])
                
        elif not ban:
            hoja_aux.append([j.value for j in fila])
    
    print(f'Registro: {ind} de {tot}')
    
    ind = 0
    tot = hoja_aux.max_row

    hoja_original.delete_rows(2, hoja_original.max_row)
    print('sustituyendo hoja destino...')
    for aux in hoja_aux.iter_rows(min_row=1):
        ind = ind + 1
        print(f'Registro: {ind} de {tot}', end='\r')
        hoja_original.append([a.value for a in aux])
    
    print(f'Registro: {ind} de {tot}')


    libro_aux.close()

def codigo_programatico(hoja,col_mover):
    total_col = hoja.max_column
    partida = ' '
    dependencia = ' '
    subdependencia = ' '
    fin_pos = total_col - col_mover    # 11 - 8 = 3

    print('Organizando columnas...')
    # Mover las columnas a la derecha
    for col_index in range(total_col, fin_pos, -1):
        for row_index in range(1, hoja.max_row + 1):
            source_cell = hoja.cell(row=row_index, column=col_index)
            target_cell = hoja.cell(row=row_index, column=col_index + fin_pos)
            target_cell.value = source_cell.value

    # Borrar el contenido de las columnas originales
    for col_index in range(fin_pos+1, fin_pos+4, +1):
        for row_index in range(1, hoja.max_row + 1):
            hoja.cell(row=row_index, column=col_index).value = None

    print('Agrergando Columnas Partida, Dependencia y Subdependencia...')
    hoja['D1'].value = 'Partida'
    hoja['E1'].value = 'Dependencia'
    hoja['F1'].value = 'Subdependencia'

    ind = 0
    tot = hoja.max_row - 1

    for item in hoja.iter_rows(min_row=2):
        ind = ind + 1
        print(f'Registro: {ind} de {tot}', end='\r')
        prog_code = item[2].value
        if prog_code is not None:
            partida = prog_code[13:16]
            dependencia = prog_code[8:11]
            subdependencia = prog_code[11:13]
                
            item[3].value = partida
            item[4].value = dependencia
            item[5].value = subdependencia
    
    print(f'Registro: {ind} de {tot}')

def gen_tabla_aux_nom(dataframe):
    data_aux = {
            'Partida': [],
            'Dependencia': [],
            'Subdependencia': [],
            'Monto': [],
            'Cant. registros':[]
    }

    ind = 0


    print('Filtrando y ordenando dataframe de Nomina...')
    grouped_data = dataframe.groupby(['Partida', 'Dependencia', 'Subdependencia'])
    tot = grouped_data.ngroups
    for group, group_data in grouped_data:
            ind = ind + 1
            print(f'Registro: {ind} de {tot}', end='\r')
            suma_montos = round(group_data['amount'].apply(float).sum(),2)
            cant_registros = len(group_data)
            data_aux['Partida'].append(group[0])
            data_aux['Dependencia'].append(group[1])
            data_aux['Subdependencia'].append(group[2])
            data_aux['Monto'].append(suma_montos)
            data_aux['Cant. registros'].append(cant_registros)
    
    print(f'Registro: {ind} de {tot}')

        
    dataframe_aux = pd.DataFrame(data_aux)
    return dataframe_aux.sort_values(by=['Partida', 'Dependencia', 'Subdependencia'])

def gen_tabla_aux_siif(dataframe):
    data_aux = {
            'Partida': [],
            'Dependencia': [],
            'Subdependencia': [],
            'Monto': [],
            'Cant. registros':[],
            'Código Programático': [],
            'Cuenta': []
    }

    ind = 0

    print('Filtrando y ordenando dataframe de SIIF...')
    filter_data = dataframe[dataframe['Cuenta'].str[0] == '5']
    grouped_data = filter_data.groupby(dataframe['Código Programático'].str[8:11] + dataframe['Código Programático'].str[11:13] + dataframe['Código Programático'].str[13:16])
    tot = grouped_data.ngroups
    for group, group_data in grouped_data:
            ind = ind + 1
            print(f'Registro: {ind} de {tot}', end='\r')
            suma_montos = group_data['Debe'].sum()
            cant_registros = len(group_data)
            data_aux['Partida'].append(int(group_data['Código Programático'].iloc[0][13:16]))
            data_aux['Dependencia'].append(int(group_data['Código Programático'].iloc[0][8:11]))
            data_aux['Subdependencia'].append(int(group_data['Código Programático'].iloc[0][11:13]))
            data_aux['Monto'].append(suma_montos)
            data_aux['Cant. registros'].append(cant_registros)
            data_aux['Código Programático'].append(group_data['Código Programático'].iloc[0])
            data_aux['Cuenta'].append(group_data['Cuenta'].iloc[0])
    
    print(f'Registro: {ind} de {tot}')

        
        
    dataframe_aux = pd.DataFrame(data_aux)
    return dataframe_aux.sort_values(by=['Partida', 'Dependencia', 'Subdependencia'])

def comparar(df_nomina,df_siif):

    print('Comparando...')
    merged_df = pd.merge(df_nomina, df_siif, on=['Partida', 'Dependencia', 'Subdependencia'], suffixes=('_nomina', '_siif'))

    print('Calculando la diferencia de los montos...')
    # Calcula la diferencia de los montos y almacenarla en una nueva columna "diferencia"
    merged_df['Diferencia'] = round(merged_df['Monto_nomina'] - merged_df['Monto_siif'],2)

    print('Creando Presentacion de los datos...')
    # Crear un nuevo DataFrame con las columnas 
    df_comparacion = merged_df

    with pd.ExcelWriter('Comparación de Polizas de percepciones.xlsx') as writer:
        df_comparacion.to_excel(writer, sheet_name='Comparación', index=False)
        df_nomina.to_excel(writer,sheet_name='auxiliar_nomina', index=False)
        df_siif.to_excel(writer,sheet_name='auxiliar_siif', index=False)

def start_comparar():
    if ruta3 and ruta2:
        try:
            t_inicio = time.time()
            # Cargar los archivos de Excel
            print(f'Abriendo archivo: {ruta2}')
            data_nomina = pd.read_excel(ruta2,sheet_name=0)
            print(f'Abriendo archivo: {ruta3}')
            data_siif = pd.read_excel(ruta3, sheet_name=0)  # se posiciona en la hoja 1 independientemente de como se llame

            aux_nom = gen_tabla_aux_nom(data_nomina)
            aux_siif = gen_tabla_aux_siif(data_siif)

            comparar(aux_nom, aux_siif)
            t_final = time.time()
            tiempo = t_final - t_inicio
            horas = int(tiempo / 3600)
            minutos = int((tiempo % 3600) / 60)
            segundos = int(tiempo % 60)

            mostrar_ventana_emergente(f'Comparación Finalizada, Tiempo de ejecucion {horas}hr:{minutos}min:{segundos}seg')
        except Exception as e:

            # Manejar cualquier error que pueda ocurrir al leer el archivo
            print("Error al cargar el archivo:", e)
    else:
        mostrar_ventana_emergente('Falta alguno de estos archivos: Archivo Original o Archivo de Poliza')



ventana = tk.Tk()
ventana.geometry("400x600")
ventana.title("Ajustes de Percepciones")

etiqueta = tk.Label(ventana, text='-------------Paso 1---------------')
etiqueta.pack()

boton_arch1 = tk.Button(ventana, text="Cargar archivo de Ajustes", command=cargar_archivo1)
dir_path1 = tk.Label(ventana, text='')
boton_arch2 = tk.Button(ventana, text="Cargar archivo original", command=cargar_archivo2)
dir_path2 = tk.Label(ventana, text='')
boton_arch1.pack(pady=5)
dir_path1.pack(pady=5)
boton_arch2.pack(pady=5)
dir_path2.pack(pady=5)

boton_start = tk.Button(ventana, text="Ajustar", command=cargar_archivo_excel)
boton_start.pack(pady=5)

tabla = ttk.Treeview(ventana)
tabla.pack(padx=10, pady=5)

etiqueta = tk.Label(ventana, text='----------------Paso 2------------------')
etiqueta.pack()

boton_arch3 = tk.Button(ventana, text="Cargar achivo de poliza SIIF", command=cargar_archivo3)
boton_arch3.pack(pady=5)
dir_path3 = tk.Label(ventana, text='')
dir_path3.pack(pady=5)

boton_comp = tk.Button(ventana, text="Comparar archivos", command=start_comparar)
boton_comp.pack(pady=5)

ventana.mainloop()
