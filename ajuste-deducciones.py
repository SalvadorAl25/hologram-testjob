from openpyxl import workbook
import tkinter as tk
from tkinter import filedialog, ttk
import openpyxl
import pandas as pd
import time

ruta1 = ''
ruta2 = ''
ruta3 = ''
ruta4 = ''

def mostrar_ventana_eme_msj(msj):
    ventana_emergente = tk.Toplevel(ventana)
    ventana_emergente.title("Finalizado")
    ventana_emergente.geometry("800x100")
    
    etiqueta = tk.Label(ventana_emergente, text=msj)
    etiqueta.pack(pady=20)
    
    boton_cerrar = tk.Button(ventana_emergente, text="Cerrar ventana", command=ventana_emergente.destroy)
    boton_cerrar.pack(pady=10)

def cargar_archivo1():
    global ruta1 
    ruta1= filedialog.askopenfilename(title="Seleccionar archivo de ajustes", filetypes=(("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")))
    if ruta1:
        dir_path1.config(text=f"Archivo 1 seleccionado: {ruta1}")

def cargar_archivo2():
    global ruta2 
    ruta2 = filedialog.askopenfilename(title="Seleccionar archivo original", filetypes=(("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")))
    if ruta2:
        dir_path2.config(text=f"Archivo 2 seleccionado: {ruta2}")

def cargar_archivo3():
    global ruta3 
    ruta3 = filedialog.askopenfilename(title="Seleccionar archivo De Catálogo", filetypes=(("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")))
    if ruta3:
        dir_path3.config(text=f"Archivo 3 seleccionado: {ruta3}")

def cargar_archivo4():
    global ruta4 
    ruta4 = filedialog.askopenfilename(title="Seleccionar archivo de Poliza SIIF", filetypes=(("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")))
    if ruta4:
        dir_path4.config(text=f"Archivo 3 seleccionado: {ruta4}")

def mostrar_tabla(hoja):
    # Limpiar la tabla si ya existe
    for widget in tabla.winfo_children():
        widget.destroy()

    # Obtener los títulos de las columnas
    columnas = [cell.value for cell in hoja[1]]

    # Crear la tabla usando un widget Treeview de Tkinter
    tabla["columns"] = columnas
    tabla.heading("#0", text="Índice")
    tabla.column("#0", width=50)
    for col in columnas:
        tabla.heading(col, text=col)
        tabla.column(col, width=100)

    # Obtener los datos de las filas y mostrarlos en la tabla
    for idx, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=1):
        tabla.insert("", tk.END, text=str(idx), values=fila)

def rfc_asign(hoja):
    rfc_aux = ' '
    
    for fila in hoja.iter_rows(min_row=2):
        if fila[0].value != None:
            rfc_aux = fila[0].value
        else:
            fila[0].value = rfc_aux
    print(f"rfc asignados...")
         
def buscar_per_rfc(rfc, hoja_ajustes):
    busqueda = []
    for fila in hoja_ajustes.iter_rows(min_row=2):
        if fila[0].value == rfc:
            busqueda.append(fila)
    return busqueda

def transformar(hoja_ajustes,hoja_original):
    rfc_aux = ' '
    busqueda = []
    ban = False
    # Crear un nuevo archivo de Excel
    # este libro es auxiliar, aqui se iran concatenando los registros, para despues sobre escribir la hora destino
    libro_aux = openpyxl.Workbook()
    hoja_aux = libro_aux.active

    ind = 0
    tot = hoja_original.max_row - 1

    print(f"transformando... ")

    for fila in hoja_original.iter_rows(min_row=2):
        ind = ind + 1
        print(f'Registro: {ind} de {tot}', end='\r')
        if fila[0].value != rfc_aux:
            rfc_aux = fila[0].value
            ban = False
            busqueda = buscar_per_rfc(rfc_aux,hoja_ajustes)
            if len(busqueda) != 0:
                for b in busqueda:
                    hoja_aux.append([i.value for i in b])
                    ban = True
            else:
                hoja_aux.append([j.value for j in fila])
                
        elif not ban:
            hoja_aux.append([j.value for j in fila])

    print(f'Registro: {ind} de {tot}')

    ind = 0
    tot = hoja_aux.max_row

    print('sustituyendo hoja destino...')
    hoja_original.delete_rows(2, hoja_original.max_row)
    for aux in hoja_aux.iter_rows(min_row=1):
        ind = ind + 1
        print(f'Registro: {ind} de {tot}', end='\r')
        hoja_original.append([a.value for a in aux])
    
    print(f'Registro: {ind} de {tot}')

    libro_aux.close()

def clv_deduccion(hoja_deducciones, hoja_cat_ded):
    cuenta = ' '
    hoja_deducciones['F1'] = 'Cuenta Contable SIIF'
    tot = hoja_deducciones.max_row - 1
    ind = 0

    print('Obteniendo Cuenta Contable SIIF...')

    for fila_ded in hoja_deducciones.iter_rows(min_row=2):
        ind = ind + 1
        print(f'Registro: {ind} de {tot}', end='\r')
        for fila_cat in hoja_cat_ded.iter_rows(min_row=2):
            if fila_ded[2].value == fila_cat[0].value:
                cuenta = fila_cat[2].value
                fila_ded[5].value = cuenta[:15]
    
    print(f'Registro: {ind} de {tot}')
    
def cargar_archivo_excel():
    if ruta1 and ruta2 and ruta3:
        try:
            t_inicio = time.time()
            print(f'Cargando archivo {ruta1}')
            libro_ajustes = openpyxl.load_workbook(ruta1)
            print(f'Cargando archivo {ruta2}')
            libro_original = openpyxl.load_workbook(ruta2)
            print(f'Cargando archivo {ruta3}')
            libro_catalogo = openpyxl.load_workbook(ruta3)

            hoja_ajustes = libro_ajustes.active
            hoja_original = libro_original.active
            hoja_catalogo = libro_catalogo.active

            rfc_asign(hoja_ajustes)
            rfc_asign(hoja_original)
            transformar(hoja_ajustes,hoja_original)
            clv_deduccion(hoja_original,hoja_catalogo)
            print('Guardando Archivo...')
            libro_original.save(ruta2)
            print('Cerrando Archivo...')
            libro_original.close()
            # Mostrar el contenido en la tabla
            mostrar_tabla(hoja_original)
            libro_ajustes.close()
            libro_catalogo.close()
            t_final = time.time()
            tiempo = t_final - t_inicio
            horas = int(tiempo / 3600)
            minutos = int((tiempo % 3600) / 60)
            segundos = int(tiempo % 60)
            mostrar_ventana_eme_msj(f'Transformación Finalizada, Tiempo de ejecucion {horas}hr:{minutos}min:{segundos}seg')
        except Exception as e:
            # Manejar cualquier error que pueda ocurrir al leer el archivo
            print("Error al cargar el archivo:", e)
    else:
        mostrar_ventana_eme_msj('Falta alguno de los siguientes archivos: Archivo de Ajustes o Archivo original o Archivo de Catálogo')

def gen_tabla_aux_nom(dataframe):
    data_aux = {
            'Cuenta': [],
            'Monto Nómina': [],
            'Registros': []
    }
    ind = 0

    print('Filtrando y ordenando dataframe de Nomina...')
    grouped_data = dataframe.groupby(['Cuenta Contable SIIF'])
    tot = grouped_data.ngroups
    for group, group_data in grouped_data:
            ind = ind + 1
            print(f'Registro: {ind} de {tot}', end='\r')
            suma_montos = round(group_data['amount'].apply(float).sum(),2)
            cant_registros = len(group_data)
            data_aux['Cuenta'].append(group_data['Cuenta Contable SIIF'].iloc[0])
            data_aux['Monto Nómina'].append(suma_montos)
            data_aux['Registros'].append(cant_registros)

    print(f'Registro: {ind} de {tot}')
        
        
    dataframe_aux = pd.DataFrame(data_aux)
    return dataframe_aux.sort_values(by=['Cuenta'])

def ajustar_haber(row):
    if row['Haber'] == 0:
        return row['Debe'] * -1
    else:
        return row['Haber']
    
def comparar(df_nomina,df_siif):

    print('comparando....')
    merged_df = pd.merge(df_nomina, df_siif, on=['Cuenta'])

    print('Calculando la diferencia de los montos...')
    merged_df['Haber'] = merged_df.apply(ajustar_haber, axis=1)

    # Calcula la diferencia de los montos y almacenarla en una nueva columna "diferencia"
    merged_df['Diferencia'] = round(merged_df['Monto Nómina'] - round(merged_df['Haber'],2),2)

    print('Creando Presentacion de los datos...')
    # Crear un nuevo DataFrame con las columnas 
    df_comparacion = merged_df[['Cuenta','Nombre','Descripción','Registros','Monto Nómina','Haber','Diferencia']]

    df_comp = df_comparacion.rename(columns={'Haber':'Monto SIIF'})

    with pd.ExcelWriter('Comparación de Polizas de deducciones.xlsx') as writer:
        df_comp.to_excel(writer, sheet_name='Comparación', index=False)
        df_nomina.to_excel(writer,sheet_name='auxiliar_nomina', index=False)
        df_siif.to_excel(writer,sheet_name='auxiliar_siif', index=False)

def start_comparar():
    if ruta2 and ruta4:
        try:
            t_inicio = time.time()
            # Cargar los archivos de Excel
            data_nomina = pd.read_excel(ruta2,sheet_name=0)
            data_siif = pd.read_excel(ruta4, sheet_name=0)  # se posiciona en la hoja 1 independientemente de como se llame

            aux_nom = gen_tabla_aux_nom(data_nomina)

            comparar(aux_nom, data_siif)
            t_final = time.time()
            tiempo = t_final - t_inicio
            horas = int(tiempo / 3600)
            minutos = int((tiempo % 3600) / 60)
            segundos = int(tiempo % 60)
            mostrar_ventana_eme_msj(f'Comparación Finalizada, Tiempo de ejecucion {horas}hr:{minutos}min:{segundos}seg')

        except Exception as e:

            # Manejar cualquier error que pueda ocurrir al leer el archivo
            print("Error al cargar el archivo:", e)
    else:
        mostrar_ventana_eme_msj('falta alguno alguno de estos archivos: Archivo Original o Archivo de Poliza')


ventana = tk.Tk()
ventana.geometry("400x650")
ventana.title("Ajuste de Deducciones")

etiqueta = tk.Label(ventana, text='-------------Paso 1---------------')
etiqueta.pack()

boton_arch1 = tk.Button(ventana, text="Cargar archivo de Ajustes", command=cargar_archivo1)
boton_arch2 = tk.Button(ventana, text="Cargar archivo original", command=cargar_archivo2)
boton_arch3 = tk.Button(ventana, text="Cargar archivo de Catálogo", command=cargar_archivo3)
dir_path1 = tk.Label(ventana, text='')
dir_path2 = tk.Label(ventana, text='')
dir_path3 = tk.Label(ventana, text='')
boton_arch1.pack(pady=5)
dir_path1.pack(pady=5,padx=5)
boton_arch2.pack(pady=5)
dir_path2.pack(pady=5,padx=5)
boton_arch3.pack(pady=5)
dir_path3.pack(pady=5,padx=5)

boton_start = tk.Button(ventana, text="Ajustar", command=cargar_archivo_excel)
boton_start.pack(pady=5)

tabla = ttk.Treeview(ventana)
tabla.pack(padx=2, pady=5)

etiqueta = tk.Label(ventana, text='----------------Paso 2------------------')
etiqueta.pack()

boton_arch4 = tk.Button(ventana, text="Cargar achivo de poliza SIIF", command=cargar_archivo4)
boton_arch4.pack(pady=5)
dir_path4 = tk.Label(ventana, text='')
dir_path4.pack(pady=5,padx=5)

boton_comp = tk.Button(ventana, text="Comparar archivos", command=start_comparar)
boton_comp.pack(pady=5)

ventana.mainloop()
